#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
_dump-source.py —— 从 烦恼分类.xlsx 抽出机器可信的原始数据，写成 _source.json

只做抽取和结构化，不做任何创作。输出给 _gen-data.py 用来生成 js/worry-data.js
与 js/gadget-data.js。xlsx 内容有变动时重跑这一步。

用法：python assets/dev/_dump-source.py [源素材根目录]
"""

import os
import re
import sys
import json

import openpyxl

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
SRC_ROOT = os.path.abspath(sys.argv[1]) if len(sys.argv) > 1 else os.path.dirname(REPO)
XLSX = os.path.join(SRC_ROOT, '烦恼', '烦恼分类.xlsx')
OUT = os.path.join(os.path.dirname(__file__), '_source.json')

# xlsx 的大类全名 -> 界面短标签（规格文案里的 9 类名）+ 素材图片序号
CATEGORY_META = {
    'C01': ('家庭', 1),
    'C02': ('亲密关系', 2),
    'C03': ('学业', 3),
    'C04': ('工作', 4),
    'C05': ('社交', 5),
    'C06': ('生活', 6),      # 素材文件名是「日常生活」
    'C07': ('经济', 7),
    'C08': ('未来', 8),      # 素材文件名是「人生」
    'C09': ('情绪', 9),
}


def clean(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


def read_worries(wb):
    ws = wb['烦恼内容分类结果']
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    cats = {}
    order = []
    worries = []
    cur_code = cur_name = None
    for r in rows:
        # 分类编号与大类是合并单元格，各自独立向下填充
        if clean(r[0]):
            cur_code = clean(r[0])
        if clean(r[1]):
            cur_name = clean(r[1])
        text = clean(r[2])
        if not text or not cur_code:
            continue
        if cur_code not in cats:
            label, img = CATEGORY_META[cur_code]
            cats[cur_code] = {
                'id': cur_code, 'label': label, 'fullName': cur_name,
                'note': clean(r[6]), 'imageIndex': img, 'count': 0,
            }
            order.append(cur_code)
        if clean(r[6]) and not cats[cur_code]['note']:
            cats[cur_code]['note'] = clean(r[6])
        cats[cur_code]['count'] += 1
        worries.append({
            'id': int(clean(r[3])),
            'text': text,
            'category': cur_code,
            'gadget': clean(r[4]),
            'gadgetGroup': clean(r[5]),
        })
    return [cats[c] for c in order], worries


def read_gadgets(wb):
    ws = wb['道具统计']
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        group, name, desc = clean(r[0]), clean(r[1]), clean(r[4])
        if not name:
            continue
        out.append({'name': name, 'group': group, 'description': desc})
    return out


def read_summaries(wb):
    ws = wb['烦恼总结']
    out = {}
    for r in ws.iter_rows(min_row=5, values_only=True):
        text, summary = clean(r[0]), clean(r[1])
        if text and summary:
            out[text] = summary
    return out


def main():
    if not os.path.exists(XLSX):
        raise SystemExit('找不到 %s' % XLSX)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    categories, worries = read_worries(wb)
    gadgets = read_gadgets(wb)
    summaries = read_summaries(wb)

    # 把总结按烦恼文案挂到每条烦恼上，并报告缺口
    missing = []
    for w in worries:
        s = summaries.get(w['text'])
        if s is None:
            missing.append(w['text'])
            w['summary'] = ''
        else:
            w['summary'] = s
    extra = sorted(set(summaries) - {w['text'] for w in worries})

    gadget_names = {g['name'] for g in gadgets}
    unknown = sorted({w['gadget'] for w in worries} - gadget_names)
    unused = sorted(gadget_names - {w['gadget'] for w in worries})

    data = {
        'categories': categories,
        'worries': worries,
        'gadgets': gadgets,
    }
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print('写入 %s' % OUT)
    print('大类 %d 个：%s' % (len(categories), ' '.join('%s%s=%d' % (c['id'], c['label'], c['count']) for c in categories)))
    print('烦恼 %d 条，道具 %d 个' % (len(worries), len(gadgets)))
    print('缺总结：%s' % (missing or '无'))
    print('多余总结：%s' % (extra or '无'))
    print('未知道具引用：%s' % (unknown or '无'))
    print('未被引用的道具：%s' % (unused or '无'))


if __name__ == '__main__':
    main()
